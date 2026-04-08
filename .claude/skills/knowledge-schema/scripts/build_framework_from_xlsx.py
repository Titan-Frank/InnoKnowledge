#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
from collections import OrderedDict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook


def stable_id(prefix: str, *parts: str) -> str:
    text = "|".join(parts)
    digest = hashlib.sha1(text.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}:{digest}"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build a curriculum framework JSON from an Excel workbook."
    )
    parser.add_argument("--xlsx", required=True, help="Path to the source workbook.")
    parser.add_argument("--subject", default="化学", help="Subject filter.")
    parser.add_argument("--stage", default="初中", help="School stage filter.")
    parser.add_argument("--out", required=True, help="Output JSON path.")
    parser.add_argument("--pretty", action="store_true", help="Write indented JSON.")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    domains: OrderedDict[str, OrderedDict[str, list[str]]] = OrderedDict()
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 7:
            continue

        subject = (row[0] or "").strip()
        stage = (row[1] or "").strip()
        domain = (row[4] or "").strip()
        topic = (row[5] or "").strip()
        text = (row[6] or "").strip()

        if subject != args.subject or stage != args.stage:
            continue
        if not domain or not topic or not text:
            continue

        domains.setdefault(domain, OrderedDict())
        domains[domain].setdefault(topic, [])
        domains[domain][topic].append(text)

    payload = {
        "framework_id": "framework:cn-chem-junior",
        "title": "初中化学课程参考框架",
        "source_path": str(xlsx_path),
        "generated_at": datetime.now(UTC).isoformat(),
        "filters": {
          "subject": args.subject,
          "school_stage": args.stage
        },
        "domains": [],
    }

    for domain_title, topics in domains.items():
        domain_id = stable_id("framework:cn-chem-junior:domain", domain_title)
        domain_record = {
            "id": domain_id,
            "title": domain_title,
            "topics": [],
        }
        for topic_title, expectations in topics.items():
            topic_id = stable_id(
                "framework:cn-chem-junior:topic",
                domain_title,
                topic_title,
            )
            topic_record = {
                "id": topic_id,
                "title": topic_title,
                "expectations": [],
            }
            for idx, text in enumerate(expectations, start=1):
                expectation_id = stable_id(
                    "framework:cn-chem-junior:expectation",
                    domain_title,
                    topic_title,
                    str(idx),
                    text,
                )
                topic_record["expectations"].append(
                    {
                        "id": expectation_id,
                        "text": text,
                    }
                )
            domain_record["topics"].append(topic_record)
        payload["domains"].append(domain_record)

    out_path = Path(args.out).expanduser()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2 if args.pretty else None)
        handle.write("\n")

    print(f"Wrote framework with {len(payload['domains'])} domains to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
